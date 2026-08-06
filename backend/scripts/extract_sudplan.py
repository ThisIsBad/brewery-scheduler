"""Extract a Sudplanung workbook (any year) into JSON for the seed import.

Two source formats, auto-detected:

- Langformat (sheet "Planung YYYY", 2025/2026): one row per Sud with a
  dated tank chain. Column layout differs per year, so columns are
  resolved by header name. Column "Biersorte" is authoritative for the
  beer (the workbooks' own summary sheets count by it).
- Kurzform (sheet "Tabelle1", the 2021-2024 log): three side-by-side
  blocks of (Sudnr., Datum, Sorte, Tank Name, Vermerk) — no tank chains,
  free-text Sorten. Normalised here to the canonical names; entries are
  marked "kurzform" so the loader imports them as finished history.

The JSON stays faithful otherwise: raw tank names (incl. the pseudo
targets "Fass"/"Ausschank"), raw dates, no sanitizing — interpretation
happens in the backend loader.

Usage: python extract_sudplan.py <Sudplanung.xlsx> [out.json]
"""

import json
import sys
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

USAGE = "Usage: python extract_sudplan.py <Sudplanung.xlsx> [out.json]"
SRC = sys.argv[1] if len(sys.argv) > 1 else sys.exit(USAGE)
OUT = sys.argv[2] if len(sys.argv) > 2 else "sudplan.json"

# Freitext-Sorten des 2021-2024-Logs -> kanonische Namen. Weicht das
# Original ab, wandert es als "Log-Sorte: ..." in die Notiz.
KURZFORM_SORTEN = {
    "keller hell": "Kellerbier Hell (Brudi)",
    "keller hell 2.0": "Kellerbier Hell (Brudi)",
    "keller hell n2": "Kellerbier Hell (Brudi)",
    "keller keller": "Kellerbier Hell (Brudi)",
    "bergbier": "Bergbier (Gisela)",
    "bergbier n3": "Bergbier (Gisela)",
    "festbier": "Bergbier (Gisela)",
    "keller bern": "Keller Bern",
    "bock": "Bock",
    "weizen": "Weizen (Fritz)",
    "weizenbock": "Weizenbock (Justus)",
    "spezial": "Spezialsud (Schwesti)",
    "spezialfantasia": "Spezialsud (Schwesti)",
    "spezial mix": "Spezialsud (Schwesti)",
    # Sud 79: Sorte im Log leer - Spezialsud als am wenigsten falsche
    # Einordnung, mit Pruef-Notiz.
    "": "Spezialsud (Schwesti)",
    "rauchbier": "Rauchbier (Waltraut)",
    "rauch n6": "Rauchbier (Waltraut)",
    "keller hell sven": "Kellerbier Hell (Sven)",
    "sven": "Kellerbier Hell (Sven)",
    "sven n4": "Kellerbier Hell (Sven)",
    "sven/keller": "Kellerbier Hell (Sven)",
    "bay. dunkel": "Bay. Dunkel (Enno)",
    "leicht rot": "Leicht Rot (Werner)",
    "orca": "Collab Sud 2024 (Orca Brau, Wit)",
}

GAERTANKS = {
    "lisa": "Lisa", "wanda": "Wanda", "greta": "Greta", "anouk": "Anouk",
    "yuri": "Yuri", "alva": "Alva", "lovis": "Lovis",
    "offener gärbottich": "Offener Gärbottich",
}


def iso(v):
    return v.date().isoformat() if isinstance(v, datetime) else None


def num(v, lo, hi):
    if isinstance(v, (int, float)) and lo <= v <= hi:
        return round(float(v), 3)
    return None


def extract_langformat(wb, blattname: str) -> list[dict]:
    ws = wb[blattname]
    header = {c.value: c.column_letter for c in ws[1] if isinstance(c.value, str)}
    col = header.get
    tank_cols = []
    for n in range(1, 6):
        t = col(f"Tank {n}")
        if t:
            tank_cols.append((t, col(f"Tankwechsel {n}")))
    nr_c, brew_c, sorte_c = col("Sud-Nr"), col("Braudatum"), col("Biersorte")
    menge_c = col("gebraute Menge [hl]")
    stw_c = col("Stammwürze [%]")
    code_c = col("Sudname")
    ausschank_c = col("Ausschank")
    versteuert_c = col("Ausschank-menge [hl] versteuert") or col("Ausschank-menge [hl]")
    header_cols = set(header.values())

    entries = []
    for row in ws.iter_rows(min_row=2):
        c = {cell.column_letter: cell.value for cell in row}
        glob = c.get(nr_c)
        if not isinstance(glob, int) or iso(c.get(brew_c)) is None:
            continue

        chain = []
        for tank_col, until_col in tank_cols:
            tank = c.get(tank_col)
            if not isinstance(tank, str) or not tank.strip():
                continue
            until = iso(c.get(until_col)) if until_col else None
            chain.append({"tank": tank.strip(), "bis": until})

        menge = None
        braumeister = None
        if str(glob) in wb.sheetnames:
            ps = wb[str(glob)]
            menge = num(ps["B7"].value, 5, 35)
            if isinstance(ps["B5"].value, str) and ps["B5"].value.strip():
                braumeister = ps["B5"].value.strip()

        # Headerlose Text-Spalten rechts sind handschriftliche Notizen.
        freitexte = [
            str(v).strip()
            for k, v in c.items()
            if k not in header_cols and isinstance(v, str) and v.strip()
        ]
        code = c.get(code_c) if code_c else None
        entries.append(
            {
                "global": glob,
                "code": code,
                "brew": iso(c[brew_c]),
                "sorte": (c.get(sorte_c) or "").strip(),
                "menge_hl": menge if menge is not None else num(c.get(menge_c), 5, 35),
                "stw": num(c.get(stw_c), 0.05, 0.25),
                "chain": chain,
                "ausschank": iso(c.get(ausschank_c)) if ausschank_c else None,
                "versteuert": num(c.get(versteuert_c), 0, 40) if versteuert_c else None,
                "braumeister": braumeister,
                "note": " · ".join(freitexte) if freitexte else None,
            }
        )
    return entries


def extract_kurzformat(wb) -> list[dict]:
    ws = wb["Tabelle1"]
    header = {c.column_letter: c.value for c in ws[1]}
    starts = [col for col, v in header.items() if v == "Sudnr."]

    entries = []
    for start in starts:
        i = column_index_from_string(start)
        cols = [get_column_letter(i + k) for k in range(5)]
        for row in ws.iter_rows(min_row=2):
            c = {cell.column_letter: cell.value for cell in row}
            glob, datum = c.get(cols[0]), c.get(cols[1])
            if not isinstance(glob, int) or iso(datum) is None:
                continue
            roh_sorte = str(c.get(cols[2]) or "").strip()
            sorte = KURZFORM_SORTEN.get(roh_sorte.lower())
            if sorte is None:
                raise SystemExit(f"Unbekannte Kurzform-Sorte {roh_sorte!r} (Sud {glob})")
            roh_tank = str(c.get(cols[3]) or "").strip()
            # "anouk Sven" u. ae.: erstes Wort ist der Tank.
            tank = GAERTANKS.get(roh_tank.lower()) or GAERTANKS.get(
                roh_tank.split()[0].lower() if roh_tank else ""
            )
            noten = []
            if roh_sorte and roh_sorte != sorte:
                noten.append(f"Log-Sorte: {roh_sorte}")
            if not roh_sorte:
                noten.append("Sorte im Log leer — bitte prüfen")
            entries.append(
                {
                    "global": glob,
                    "code": None,
                    "brew": iso(datum),
                    "sorte": sorte,
                    "menge_hl": None,
                    "stw": None,
                    "chain": [{"tank": tank, "bis": None}] if tank else [],
                    "ausschank": None,
                    "versteuert": None,
                    "braumeister": None,
                    "note": " · ".join(noten) if noten else None,
                    "kurzform": True,
                }
            )
    entries.sort(key=lambda e: e["global"])
    return entries


def main() -> None:
    wb = load_workbook(SRC, data_only=True)
    plan = next((n for n in wb.sheetnames if n.startswith("Planung")), None)
    if plan:
        entries = extract_langformat(wb, plan)
    elif "Tabelle1" in wb.sheetnames:
        entries = extract_kurzformat(wb)
    else:
        raise SystemExit(f"Kein bekanntes Format in {SRC}")

    with open(OUT, "w") as f:
        json.dump(entries, f, ensure_ascii=False, indent=1)
    print(f"{len(entries)} Sude -> {OUT}")
    styles: dict[str, int] = {}
    for e in entries:
        styles[e["sorte"]] = styles.get(e["sorte"], 0) + 1
    for s, n in sorted(styles.items(), key=lambda kv: -kv[1]):
        print(f"  {n:3d}  {s}")


if __name__ == "__main__":
    main()
