"""Extract Vincenz' Sudplanung 2026 into JSON for the seed import.

Reads sheet "Planung 2026" (one row per Sud, global numbers 210-300) plus
each Sud's protocol sheet (actual volume, Braumeister). Column F (Biersorte)
is authoritative for the beer — the workbook's own Sudübersicht counts by it
(34 Bergbier, 38 Keller inkl. Sven only add up via F); the slot letters in
column B are planning relics and are kept as "code" for reference only.

The JSON stays faithful to the sheet: raw tank names (incl. the pseudo
targets "Fass"/"Ausschank"), raw dates, no sanitizing. Interpretation
(mapping, chain repair, statuses) happens in the backend loader.
"""

import json
import sys
from datetime import datetime

from openpyxl import load_workbook

USAGE = "Usage: python extract_sudplan_2026.py <2026_Sudplanung.xlsx> [out.json]"
SRC = sys.argv[1] if len(sys.argv) > 1 else sys.exit(USAGE)
OUT = sys.argv[2] if len(sys.argv) > 2 else "sudplan_2026.json"


def iso(v):
    return v.date().isoformat() if isinstance(v, datetime) else None


def num(v, lo, hi):
    if isinstance(v, (int, float)) and lo <= v <= hi:
        return round(float(v), 3)
    return None


def main() -> None:
    wb = load_workbook(SRC, data_only=True)
    ws = wb["Planung 2026"]
    entries = []
    for row in ws.iter_rows(min_row=3, max_row=186):
        c = {cell.column_letter: cell.value for cell in row}
        if not isinstance(c.get("A"), int) or iso(c.get("E")) is None:
            continue  # Reinigungs-Zeilen und leere Nummern-Stubs

        glob = c["A"]
        # Chain: tank G until J, tank K until L, tank M until N, tank O open.
        chain = []
        for tank_col, until_col in (("G", "J"), ("K", "L"), ("M", "N"), ("O", None)):
            tank = c.get(tank_col)
            if not isinstance(tank, str) or not tank.strip():
                continue
            until = iso(c.get(until_col)) if until_col else None
            chain.append({"tank": tank.strip(), "bis": until})

        # Protocol sheet: actual volume + Braumeister (may be missing).
        menge = None
        braumeister = None
        if str(glob) in wb.sheetnames:
            ps = wb[str(glob)]
            menge = num(ps["B7"].value, 5, 35)
            if isinstance(ps["B5"].value, str) and ps["B5"].value.strip():
                braumeister = ps["B5"].value.strip()

        note = c.get("W")
        entries.append(
            {
                "global": glob,
                "code": c.get("B"),
                "brew": iso(c["E"]),
                "sorte": (c.get("F") or "").strip(),
                "menge_hl": menge if menge is not None else num(c.get("H"), 5, 35),
                "stw": num(c.get("I"), 0.05, 0.25),
                "chain": chain,
                "ausschank": iso(c.get("P")),
                "versteuert": num(c.get("Q"), 0, 40),
                "braumeister": braumeister,
                "note": note.strip() if isinstance(note, str) and note.strip() else None,
            }
        )

    with open(OUT, "w") as f:
        json.dump(entries, f, ensure_ascii=False, indent=1)
    print(f"{len(entries)} Sude -> {OUT}")
    styles = {}
    for e in entries:
        styles[e["sorte"]] = styles.get(e["sorte"], 0) + 1
    for s, n in sorted(styles.items(), key=lambda kv: -kv[1]):
        print(f"  {n:3d}  {s}")


if __name__ == "__main__":
    main()
